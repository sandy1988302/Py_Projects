import smtplib
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
import time

usernames = {}
wb = openpyxl.load_workbook('邮件列表.xlsx')
ws = wb.active
for i in range(2, ws.max_row + 1):
    usernames.setdefault(ws['A' + str(i)].value, {})
    usernames[ws['A' + str(i)].value].setdefault('电子邮件', ws['D' + str(i)].value)
    if str(ws['B' + str(i)].value) == '男':
        usernames[ws['A' + str(i)].value].setdefault('称谓', '先生')
    else:
        usernames[ws['A' + str(i)].value].setdefault('称谓', '女士')
username = '290591844@qq.com'
passwrod = 'inukdsuhfhqybhjh'
frommail = '290591844@qq.com'
subject = '*****公司的邀请函'
server = smtplib.SMTP(host='smtp.qq.com', port=25)
server.login(user=username, password=passwrod)


def body(name, male):
    texts = '''
    尊敬的{}{}:
        邀请您参加本公司于3月1日于****举办的公司展销会!'''.format(name, male)
    return texts


for key in usernames.keys():
    msg = MIMEMultipart()
    msg['From'] = frommail
    if usernames[key]['电子邮件'] is None or str(usernames[key]['电子邮件']).find('@') == -1:
        print('{}没有填写电子邮件地址!或邮箱地址不正确!')
        continue
    print(key)
    msg['To'] = usernames[key]['电子邮件']
    msg['Subject'] = '***公司给' + key + '的邀请函'
    msg.attach(MIMEText(body(key, usernames[key]['称谓'])))
    with open('亲爱的.txt', mode='rb') as f:
        attfile = f.read()
    att = MIMEApplication(attfile)  # 最重要的两行.测试其他办法添加附件会导致tcmime.1774.1903.2076.bin的问题
    att.add_header('Content-Disposition', 'attachment', filename="亲爱的.txt")
    msg.attach(att)
    try:
        server.sendmail(frommail, usernames[key]['电子邮件'],
                        msg.as_string())  # 还有这行 测试过用server.send_message QQ邮箱接受的时候还是会把附件自动改成bin文件.
    except smtplib.SMTPSenderRefused:
        print('fail')
        time.sleep(10)
    except smtplib.SMTPDataError:
        print('fail2')
        time.sleep(10)
server.close()
